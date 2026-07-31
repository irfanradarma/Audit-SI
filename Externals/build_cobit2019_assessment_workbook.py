from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

out_path = Path(r"c:\Users\USER\OneDrive - Kemenkeu\PKN STAN\Audit SI\Externals\COBIT_2019_IT_Governance_Assessment.xlsx")

objectives = [
    ("EDM01", "Ensured Governance Framework Setting and Maintenance", "Establishes and maintains the governance system, principles, and decision-making structure for enterprise IT.", "Provide clear oversight, accountability, and alignment of IT governance with enterprise strategy."),
    ("EDM02", "Ensured Benefits Delivery", "Ensures that IT-enabled investments and services deliver the expected business value and benefits.", "Maximize value from IT investments and ensure strategic business outcomes are realized."),
    ("EDM03", "Ensured Risk Optimization", "Ensures that IT-related risks are identified, assessed, and managed in line with enterprise risk appetite.", "Protect the enterprise from unacceptable IT risk while supporting innovation and growth."),
    ("EDM04", "Ensured Resource Optimization", "Ensures that IT resources are allocated and used efficiently to support enterprise priorities.", "Improve efficiency, reduce waste, and sustain the required capability for IT services."),
    ("EDM05", "Ensured Stakeholder Engagement", "Ensures that stakeholder needs, expectations, and concerns are understood and addressed through governance.", "Strengthen trust, transparency, and stakeholder participation in IT governance."),
    ("APO01", "Managed I&T Management Framework", "Establishes and maintains the overarching management framework for information and technology.", "Create a coherent structure for planning, organizing, and governing IT activities."),
    ("APO02", "Managed Strategy", "Defines the IT strategy that supports enterprise goals and business priorities.", "Align IT direction and investments with the enterprise strategy."),
    ("APO03", "Managed Enterprise Architecture", "Designs and maintains the enterprise architecture for business, information, data, applications, and technology.", "Improve integration, consistency, and long-term adaptability of the IT landscape."),
    ("APO04", "Managed Innovation", "Creates and manages innovation initiatives that improve business capabilities and services.", "Drive continuous improvement and competitiveness through technology-enabled innovation."),
    ("APO05", "Managed Portfolio", "Manages the portfolio of IT initiatives and investments to deliver business value.", "Prioritize and balance investments to support enterprise objectives."),
    ("APO06", "Managed Budget and Costs", "Plans, budgets, and controls IT spending across the enterprise.", "Ensure affordability, financial discipline, and value for money in IT."),
    ("APO07", "Managed Human Resources", "Ensures that IT human resources are available, competent, and motivated.", "Build the skills, capacity, and culture needed for effective IT delivery."),
    ("APO08", "Managed Relationships", "Manages relationships between IT and business stakeholders, partners, and users.", "Strengthen collaboration, trust, and shared accountability across the enterprise."),
    ("APO09", "Managed Service Agreements", "Establishes and maintains service agreements that define expected IT services and outcomes.", "Clarify obligations and improve delivery accountability for IT services."),
    ("APO10", "Managed Vendors", "Manages third-party and vendor relationships to ensure value and compliance.", "Reduce supplier risk and improve the quality and reliability of external services."),
    ("APO11", "Managed Quality", "Defines and manages quality standards for IT processes, services, and outputs.", "Improve consistency, reliability, and customer satisfaction in IT delivery."),
    ("APO12", "Managed Risk", "Manages IT-related risk in a coordinated and enterprise-aligned manner.", "Ensure risk is understood, mitigated, and transparent to decision-makers."),
    ("APO13", "Managed Security", "Protects information and technology assets through effective security governance and operations.", "Safeguard confidentiality, integrity, availability, and resilience of enterprise assets."),
    ("BAI01", "Managed Programs", "Coordinates programs and major change initiatives to achieve business outcomes.", "Deliver strategic change and business transformation in a controlled manner."),
    ("BAI02", "Managed Requirements Definition", "Ensures that business requirements are clearly defined, agreed, and translated into solution needs.", "Improve solution quality by aligning requirements with business needs."),
    ("BAI03", "Managed Solutions Identification and Build", "Manages the identification, design, and construction of solutions that meet requirements.", "Deliver effective and fit-for-purpose solutions that support business goals."),
    ("BAI04", "Managed Availability and Capacity", "Ensures that IT services and infrastructure are available and fit for demand.", "Maintain business continuity and service quality under changing workloads."),
    ("BAI05", "Managed Organizational Change Enablement", "Manages change impacts, communication, and user adoption for IT-enabled changes.", "Increase adoption and reduce resistance to change across the organization."),
    ("BAI06", "Managed IT Changes", "Controls change implementation to reduce disruption and operational risk.", "Improve stability and predictability of IT services during change."),
    ("BAI07", "Managed IT Change Acceptance and Transitioning", "Manages acceptance testing, transition, and release into production.", "Ensure new or changed services are fit for use and can be supported effectively."),
    ("BAI08", "Managed Knowledge", "Captures, shares, and applies knowledge that supports IT and business needs.", "Strengthen decision-making through the effective use of knowledge and lessons learned."),
    ("BAI09", "Managed Assets", "Manages the lifecycle of IT assets and related information assets.", "Protect value and reduce losses associated with IT assets."),
    ("BAI10", "Managed Configuration", "Maintains configuration information for IT assets and services.", "Improve traceability, control, and reliability of IT environments."),
    ("BAI11", "Managed Projects", "Manages IT projects from initiation through delivery and closure.", "Increase the likelihood of successful project outcomes on time and within budget."),
    ("DSS01", "Managed Operations", "Ensures that IT operations are performed reliably and efficiently.", "Provide stable, resilient, and consistent operational services."),
    ("DSS02", "Managed Service Requests and Incidents", "Manages service requests, incidents, and restoration activities.", "Restore service quickly and maintain user confidence in IT support."),
    ("DSS03", "Managed Problems", "Identifies and resolves recurring operational issues and root causes.", "Reduce the frequency and impact of recurring service problems."),
    ("DSS04", "Managed Continuity", "Ensures that business continuity and disaster recovery arrangements are in place.", "Protect enterprise operations against disruptions and major incidents."),
    ("DSS05", "Managed Security Services", "Delivers security services and monitoring to protect the enterprise.", "Reduce exposure to cyber and information security threats."),
    ("DSS06", "Managed Business Process Controls", "Ensures that IT-supported business processes are controlled and monitored.", "Improve reliability, compliance, and integrity of business processes."),
    ("MEA01", "Managed Performance and Conformance Monitoring", "Monitors the performance of IT processes and conformance to standards and policies.", "Provide evidence of performance, defects, and improvement opportunities."),
    ("MEA02", "Managed System of Internal Control", "Evaluates whether the control environment operates effectively.", "Provide assurance that controls are adequate and functioning."),
    ("MEA03", "Managed Compliance With External Requirements", "Assesses compliance with laws, regulations, contractual obligations, and industry standards.", "Reduce legal, regulatory, and contractual exposure for the enterprise."),
    ("MEA04", "Managed Assurance", "Provides independent assurance over governance, control, and risk management activities.", "Increase confidence in the effectiveness of the governance and control system."),
]

practice_catalog = {
    "EDM01": [
        ("EDM01.01", "Governance Framework Definition", "Define governance principles, scope, and decision rights for the enterprise IT governance system.", "% of governance policies approved; number of governance forums convened; policy review cycle adherence"),
        ("EDM01.02", "Governance Oversight and Direction", "Establish board and executive oversight mechanisms and direct management on strategic priorities.", "Board review meeting frequency; % of strategic decisions documented; overdue action items"),
        ("EDM01.03", "Governance Performance Monitoring", "Monitor governance outcomes, compliance, and continuous improvement of the governance system.", "Governance KPI coverage; % of corrective actions closed; management report timeliness"),
    ],
    "EDM02": [
        ("EDM02.01", "Benefits Identification", "Identify and document expected benefits, business outcomes, and value realization targets.", "Benefit register coverage; % of benefits with owners; benefit target completeness"),
        ("EDM02.02", "Benefits Delivery Tracking", "Track value realization, monitor progress, and intervene when expected benefits are at risk.", "Benefit realization rate; variance from plan; benefit review frequency"),
        ("EDM02.03", "Benefits Review and Improvement", "Review realized benefits and improve investment governance based on lessons learned.", "Post-implementation review completion; lessons learned action rate; realized value %"),
    ],
    "EDM03": [
        ("EDM03.01", "Risk Appetite and Governance", "Set risk appetite, tolerance levels, and governance expectations for IT-related risks.", "Risk appetite statements approved; risk threshold breaches; risk review cadence"),
        ("EDM03.02", "Risk Assessment and Response", "Assess, prioritize, and respond to enterprise IT risks using structured methods.", "Risk assessments completed; high-risk actions closed; residual risk review coverage"),
        ("EDM03.03", "Risk Monitoring and Reporting", "Monitor risk trends, escalate significant issues, and report risk to governance bodies.", "Risk incident trend; management reporting timeliness; risk action closure rate"),
    ],
    "EDM04": [
        ("EDM04.01", "Resource Planning", "Plan IT resources required to support enterprise objectives and strategy.", "Resource forecast accuracy; staffing gap rate; capacity utilization"),
        ("EDM04.02", "Resource Allocation and Control", "Allocate resources to initiatives and ensure appropriate investment discipline.", "Budget variance; resource utilization; approved project staffing rate"),
        ("EDM04.03", "Resource Optimization Review", "Review efficiency, reuse, and optimization opportunities for IT resources.", "Cost per service; redundant tool ratio; optimization initiative count"),
    ],
    "EDM05": [
        ("EDM05.01", "Stakeholder Needs Identification", "Identify stakeholder needs, expectations, and concerns that affect IT governance.", "Stakeholder mapping coverage; satisfaction survey response rate; issue log completeness"),
        ("EDM05.02", "Stakeholder Communication", "Communicate governance decisions, expectations, and outcomes to relevant stakeholders.", "Communication plan completion; stakeholder awareness score; issue resolution cycle time"),
        ("EDM05.03", "Stakeholder Feedback and Trust", "Collect feedback, resolve concerns, and strengthen stakeholder trust in IT governance.", "Feedback response rate; trust index; recurring complaint rate"),
    ],
    "APO01": [
        ("APO01.01", "Management Framework Design", "Design the management framework that directs enterprise IT planning and execution.", "Framework coverage; policy approval rate; process owner assignment rate"),
        ("APO01.02", "Framework Governance", "Ensure the management framework remains current, controlled, and aligned with enterprise needs.", "Policy review turnaround; framework change approval rate; compliance evidence completeness"),
        ("APO01.03", "Framework Improvement", "Use performance data and assurance results to improve the management framework.", "Improvement action rate; audit findings trend; process maturity score"),
    ],
    "APO02": [
        ("APO02.01", "Strategy Definition", "Define enterprise-aligned IT strategies, principles, and target outcomes.", "Strategy approval rate; alignment score; roadmap completeness"),
        ("APO02.02", "Strategy Translation", "Translate strategy into initiatives, priorities, and execution plans.", "Initiative-to-strategy alignment; milestone attainment; strategic KPI coverage"),
        ("APO02.03", "Strategy Review", "Review the strategy periodically and adjust based on business and technology change.", "Strategy review frequency; change requests approved; alignment variance"),
    ],
    "APO03": [
        ("APO03.01", "Architecture Blueprinting", "Develop and maintain enterprise architecture blueprints and standards.", "Architecture coverage; standard adoption rate; design review completion"),
        ("APO03.02", "Architecture Governance", "Govern architectural decisions, traceability, and compliance with standards.", "Architecture decision log; exception rate; compliance audit pass rate"),
        ("APO03.03", "Architecture Evolution", "Review and update the architecture to support innovation and change.", "Architecture refresh cycle; modernization backlog; standard deviation"),
    ],
    "APO04": [
        ("APO04.01", "Innovation Planning", "Identify innovation opportunities that support enterprise value and strategic objectives.", "Innovation pipeline size; initiative prioritization coverage; business value estimate rate"),
        ("APO04.02", "Innovation Delivery", "Manage innovation initiatives from concept through implementation and adoption.", "Pilot-to-production conversion rate; adoption rate; time to value"),
        ("APO04.03", "Innovation Learning", "Capture lessons learned and incorporate improvements into future innovation efforts.", "Post-implementation review completion; improvement action rate; innovation ROI"),
    ],
    "APO05": [
        ("APO05.01", "Portfolio Definition", "Define and maintain the portfolio of IT initiatives and investments.", "Portfolio completeness; business case coverage; prioritization consistency"),
        ("APO05.02", "Portfolio Governance", "Govern investment decisions, dependencies, and resource allocation across the portfolio.", "Decision log completeness; dependency issue rate; budget adherence"),
        ("APO05.03", "Portfolio Review", "Review portfolio performance and re-balance investments based on outcomes and risks.", "Portfolio review frequency; benefit realization variance; cancelled initiative rate"),
    ],
    "APO06": [
        ("APO06.01", "Budget Planning", "Prepare and approve IT budgets that reflect business priorities and expected demand.", "Budget accuracy; planning cycle timeliness; forecast completeness"),
        ("APO06.02", "Cost Control", "Monitor actual spend, manage variances, and enforce financial discipline.", "Cost variance; expense approval compliance; budget reforecast frequency"),
        ("APO06.03", "Cost Optimization", "Identify opportunities to reduce unnecessary cost while maintaining required service levels.", "Cost optimization savings; service level stability; vendor renegotiation count"),
    ],
    "APO07": [
        ("APO07.01", "Capability Planning", "Plan and secure the skills, roles, and competencies needed for IT delivery.", "Skills gap rate; staffing coverage; training completion rate"),
        ("APO07.02", "Workforce Management", "Assign, develop, and retain IT personnel according to enterprise needs.", "Retention rate; performance review completion; role coverage"),
        ("APO07.03", "Workforce Improvement", "Improve performance through training, knowledge sharing, and capability development.", "Training hours per employee; certification rate; productivity improvement"),
    ],
    "APO08": [
        ("APO08.01", "Relationship Management", "Establish governance structures for relationships with business stakeholders and partners.", "Relationship review frequency; stakeholder satisfaction; escalations closed"),
        ("APO08.02", "Service Communication", "Ensure that expectations, commitments, and concerns are communicated clearly.", "Communication plan coverage; SLA adherence; issue resolution time"),
        ("APO08.03", "Relationship Improvement", "Monitor relationship health and improve collaboration where needed.", "Relationship score; recurring complaints; action closure rate"),
    ],
    "APO09": [
        ("APO09.01", "Service Agreement Definition", "Define service levels, responsibilities, and outcomes in formal agreements.", "Agreement coverage; SLA completeness; sign-off rate"),
        ("APO09.02", "Agreement Governance", "Review and control service agreements to ensure value and compliance.", "Review cycle adherence; exception rate; renewal timeliness"),
        ("APO09.03", "Agreement Performance", "Track service performance and improve agreement effectiveness over time.", "SLA attainment; service complaints; improvement action rate"),
    ],
    "APO10": [
        ("APO10.01", "Vendor Selection and Governance", "Select and govern vendors according to enterprise requirements and controls.", "Vendor onboarding compliance; contract coverage; due diligence completion"),
        ("APO10.02", "Vendor Performance Management", "Monitor vendor performance, risks, and service quality against expectations.", "Vendor scorecard usage; service incident rate; SLA attainment"),
        ("APO10.03", "Vendor Relationship Improvement", "Address vendor issues and improve supplier value delivery.", "Improvement action closure; dispute rate; renewal satisfaction"),
    ],
    "APO11": [
        ("APO11.01", "Quality Standards Definition", "Define quality standards and acceptance criteria for IT processes and services.", "Standard coverage; acceptance criteria completeness; quality policy approval"),
        ("APO11.02", "Quality Assurance", "Perform assurance activities to verify compliance with quality requirements.", "Quality review coverage; defect rate; audit pass rate"),
        ("APO11.03", "Quality Improvement", "Use quality findings to improve IT processes and service outcomes.", "Defects per release; improvement action closure; customer satisfaction"),
    ],
    "APO12": [
        ("APO12.01", "Risk Identification", "Identify and document IT-related risk scenarios and their business impacts.", "Risk register completeness; risk owner assignment; impact coverage"),
        ("APO12.02", "Risk Treatment", "Define controls, mitigations, and monitoring responses for significant risks.", "Control effectiveness; mitigation plan completion; residual risk review"),
        ("APO12.03", "Risk Reporting", "Report risk status, trends, and issues to decision makers.", "Risk review cadence; escalation timeliness; action closure rate"),
    ],
    "APO13": [
        ("APO13.01", "Security Policy and Control", "Define security policies, standards, and control expectations.", "Policy coverage; control documentation rate; policy approval timeliness"),
        ("APO13.02", "Security Operations Management", "Operate security monitoring, incident handling, and protection activities.", "Incident response time; monitoring coverage; vulnerability closure rate"),
        ("APO13.03", "Security Improvement", "Review security performance and improve resilience against emerging threats.", "Improvement plan coverage; find-remediate cycle time; control maturity score"),
    ],
    "BAI01": [
        ("BAI01.01", "Program Scope and Governance", "Define program scope, governance, and success criteria for major change initiatives.", "Program charter completeness; governance forum attendance; scope change rate"),
        ("BAI01.02", "Program Delivery", "Coordinate delivery activities, dependencies, and resource alignment for programs.", "Milestone adherence; dependency resolution time; budget variance"),
        ("BAI01.03", "Program Review", "Review program progress and benefits realization to improve execution.", "Review frequency; corrective actions closed; benefit tracking coverage"),
    ],
    "BAI02": [
        ("BAI02.01", "Requirements Elicitation", "Gather, validate, and prioritize business requirements for solutions.", "Requirement coverage; stakeholder sign-off rate; traceability completeness"),
        ("BAI02.02", "Requirements Management", "Manage requirements changes and maintain alignment to business needs.", "Change request turnaround; requirement volatility; approval timeliness"),
        ("BAI02.03", "Requirements Quality", "Ensure that requirements are clear, testable, and fit for solution delivery.", "Requirement quality score; test case coverage; rework rate"),
    ],
    "BAI03": [
        ("BAI03.01", "Solution Design", "Design solutions that meet business requirements and architecture standards.", "Design review completion; requirement traceability; standard compliance"),
        ("BAI03.02", "Solution Build", "Build and configure solution components according to approved design.", "Build completion rate; defect escape rate; delivery timeliness"),
        ("BAI03.03", "Solution Testing", "Test, validate, and prepare solutions for release and operational use.", "Test coverage; defect density; UAT success rate"),
    ],
    "BAI04": [
        ("BAI04.01", "Capacity Planning", "Plan and maintain IT capacity to meet current and future demand.", "Capacity utilization; forecast accuracy; service degradation incidents"),
        ("BAI04.02", "Availability Management", "Monitor service availability and ensure continuity of critical services.", "Availability percentage; incident count; recovery time"),
        ("BAI04.03", "Capacity Improvement", "Review performance trends and improve capacity and resilience.", "Performance trend review; tuning actions completed; service level attainment"),
    ],
    "BAI05": [
        ("BAI05.01", "Change Impact Assessment", "Assess the impact of changes on people, processes, and operations.", "Impact assessment coverage; change readiness score; stakeholder engagement rate"),
        ("BAI05.02", "Change Adoption", "Support adoption through communication, training, and readiness activities.", "Training completion; adoption rate; support request volume"),
        ("BAI05.03", "Change Reinforcement", "Reinforce new ways of working and monitor adoption sustainability.", "Adoption survey score; reinforcement action rate; process compliance"),
    ],
    "BAI06": [
        ("BAI06.01", "Change Planning", "Plan, authorize, and schedule IT changes in a controlled way.", "Change schedule adherence; emergency change rate; approval coverage"),
        ("BAI06.02", "Change Implementation", "Implement approved changes with appropriate controls and rollback plans.", "Successful change rate; rollback rate; implementation lead time"),
        ("BAI06.03", "Change Review", "Review change outcomes and improve control over future changes.", "Post-change review completion; incident correlation; improvement action rate"),
    ],
    "BAI07": [
        ("BAI07.01", "Acceptance Planning", "Plan acceptance criteria, test scenarios, and transition activities.", "Acceptance criteria coverage; test execution rate; readiness score"),
        ("BAI07.02", "Acceptance Execution", "Perform acceptance testing and verify readiness for production.", "UAT pass rate; defect severity; transition readiness index"),
        ("BAI07.03", "Transition Governance", "Transition changes to operations with clear ownership and support.", "Transition completion; handover checklist rate; operational incident rate"),
    ],
    "BAI08": [
        ("BAI08.01", "Knowledge Capture", "Capture and maintain relevant knowledge for IT solutions and operations.", "Knowledge article coverage; update frequency; reuse rate"),
        ("BAI08.02", "Knowledge Sharing", "Share knowledge across teams and stakeholders.", "Sharing session frequency; access rate; knowledge contribution rate"),
        ("BAI08.03", "Knowledge Use", "Apply knowledge to improve decision making and service delivery.", "Decision quality score; issue recurrence rate; training effectiveness"),
    ],
    "BAI09": [
        ("BAI09.01", "Asset Inventory", "Maintain an accurate inventory of IT and information assets.", "Asset coverage; inventory accuracy; ownership assignment rate"),
        ("BAI09.02", "Asset Protection", "Protect assets throughout their lifecycle with appropriate controls.", "Control coverage; incident rate; asset loss events"),
        ("BAI09.03", "Asset Lifecycle Review", "Review asset lifecycle status and improve management decisions.", "Lifecycle review frequency; obsolete asset rate; exception closure rate"),
    ],
    "BAI10": [
        ("BAI10.01", "Configuration Baseline", "Define and maintain configuration baselines for critical systems and services.", "Baseline coverage; change traceability; version control rate"),
        ("BAI10.02", "Configuration Control", "Control configuration changes and maintain integrity of deployed components.", "Unauthorized change rate; audit compliance; rollback effectiveness"),
        ("BAI10.03", "Configuration Review", "Review configuration records and improve accuracy over time.", "Review cycle completion; discrepancy rate; corrective action rate"),
    ],
    "BAI11": [
        ("BAI11.01", "Project Governance", "Govern project initiation, planning, and execution with clear accountability.", "Project charter completion; governance meeting attendance; scope variance"),
        ("BAI11.02", "Project Delivery Control", "Monitor delivery against time, cost, quality, and benefit expectations.", "On-time delivery rate; budget variance; quality defect rate"),
        ("BAI11.03", "Project Closure", "Close projects, review outcomes, and capture lessons learned.", "Closure checklist completion; lessons learned rate; benefits tracking"),
    ],
    "DSS01": [
        ("DSS01.01", "Operational Planning", "Plan and schedule operational activities required to support stable services.", "Operational plan coverage; downtime forecast accuracy; service readiness rate"),
        ("DSS01.02", "Operational Execution", "Execute operational tasks, monitor service health, and maintain service continuity.", "Incident response time; service availability; task completion rate"),
        ("DSS01.03", "Operational Improvement", "Review operations and improve resilience, efficiency, and service quality.", "Improvement actions completed; recurring incident rate; SLA attainment"),
    ],
    "DSS02": [
        ("DSS02.01", "Service Request Handling", "Manage service requests and user interactions in a controlled manner.", "Request turnaround time; backlog aging; satisfaction score"),
        ("DSS02.02", "Incident Response", "Respond to incidents, restore service, and document resolution activities.", "MTTR; restoration success rate; incident logging completeness"),
        ("DSS02.03", "Incident Review", "Review incidents and improve prevention and response effectiveness.", "Repeat incident rate; root cause closure; improvement action rate"),
    ],
    "DSS03": [
        ("DSS03.01", "Problem Identification", "Identify recurring issues and determine contributing factors.", "Problem record count; trend analysis coverage; root cause assignment"),
        ("DSS03.02", "Problem Resolution", "Resolve problems with corrective and preventive actions.", "Corrective action closure; recurrence rate; resolution lead time"),
        ("DSS03.03", "Problem Review", "Review problem management outcomes and ensure continuous improvement.", "Review frequency; action effectiveness; issue recurrence"),
    ],
    "DSS04": [
        ("DSS04.01", "Continuity Planning", "Define continuity, recovery, and backup procedures for critical services.", "BCP coverage; recovery plan completeness; backup success rate"),
        ("DSS04.02", "Continuity Testing", "Test continuity arrangements and validate business recovery capabilities.", "Test frequency; recovery success rate; gap closure"),
        ("DSS04.03", "Continuity Improvement", "Update continuity arrangements based on lessons learned and evolving requirements.", "Improvement action rate; test results trend; resilience maturity"),
    ],
    "DSS05": [
        ("DSS05.01", "Security Monitoring", "Monitor security events, logs, and alerts in support of enterprise protection.", "Monitoring coverage; alert response time; detection effectiveness"),
        ("DSS05.02", "Security Response", "Respond to security incidents and maintain control of the environment.", "MTTR for security incidents; containment time; incident closure rate"),
        ("DSS05.03", "Security Improvement", "Review security performance and improve the control environment over time.", "Vulnerability closure rate; control maturity; lessons learned adoption"),
    ],
    "DSS06": [
        ("DSS06.01", "Control Design", "Define controls for business process automation and IT-supported operations.", "Control coverage; control owner assignment; control document completeness"),
        ("DSS06.02", "Control Monitoring", "Monitor control operation, exceptions, and remediation activities.", "Exception rate; monitoring completeness; remediation timeliness"),
        ("DSS06.03", "Control Improvement", "Improve control design and operation based on evidence and findings.", "Improvement action closure; control effectiveness; audit finding trend"),
    ],
    "MEA01": [
        ("MEA01.01", "Performance Measurement", "Define and collect performance indicators for key IT processes and services.", "KPI coverage; reporting completeness; metric quality"),
        ("MEA01.02", "Conformance Review", "Review adherence to policies, standards, and operating expectations.", "Compliance review coverage; nonconformance rate; remediation timeliness"),
        ("MEA01.03", "Performance Improvement", "Use performance and conformance data to improve process effectiveness.", "Improvement action closure; trend analysis coverage; maturity score"),
    ],
    "MEA02": [
        ("MEA02.01", "Control Environment Review", "Assess whether controls are documented, assigned, and operating as intended.", "Control inventory coverage; control owner rate; evidence availability"),
        ("MEA02.02", "Control Testing", "Test key controls and evaluate the sufficiency of monitoring activities.", "Test coverage; exception rate; control test pass rate"),
        ("MEA02.03", "Control Assurance", "Report assurance findings and support continuous improvement of the control system.", "Assurance report timeliness; findings closure; management follow-up rate"),
    ],
    "MEA03": [
        ("MEA03.01", "Compliance Scope", "Identify laws, regulations, contracts, and obligations applicable to IT.", "Obligation inventory coverage; legal review completion; policy mapping rate"),
        ("MEA03.02", "Compliance Monitoring", "Monitor compliance evidence, exceptions, and remediation actions.", "Compliance evidence completeness; issue count; remediation timeliness"),
        ("MEA03.03", "Compliance Reporting", "Report compliance status, risk exposure, and governance actions.", "Reporting frequency; issue escalation rate; management response time"),
    ],
    "MEA04": [
        ("MEA04.01", "Assurance Planning", "Plan assurance coverage across key governance, control, and risk domains.", "Assurance plan completeness; coverage rate; resource allocation"),
        ("MEA04.02", "Assurance Execution", "Perform assurance activities and validate evidence for key controls and processes.", "Evidence quality; issue detection rate; testing completion"),
        ("MEA04.03", "Assurance Reporting", "Report assurance results, recommendations, and follow-up actions to governance bodies.", "Report timeliness; recommendation closure; management follow-up rate"),
    ],
}

activity_catalog = {}
for objective_code, practices in practice_catalog.items():
    for practice_code, practice_name, practice_description, example_metrics in practices:
        activities = []
        if practice_code.endswith(".01"):
            activities = [
                ("Define scope, ownership, and success criteria for the practice.", 1),
                ("Document the required policies, procedures, and evidence.", 2),
                ("Assign accountable roles and review dates for ongoing execution.", 2),
                ("Review the practice outcome with management and governance bodies.", 3),
            ]
        elif practice_code.endswith(".02"):
            activities = [
                ("Execute the core operating activities for the practice.", 2),
                ("Monitor progress, exceptions, and required follow-up actions.", 3),
                ("Escalate issues that require governance attention.", 3),
                ("Capture evidence and maintain records for oversight.", 4),
            ]
        else:
            activities = [
                ("Measure results against the intended objective and target metrics.", 3),
                ("Identify improvement opportunities and remediation actions.", 4),
                ("Implement approved changes and update supporting documentation.", 4),
                ("Reassess effectiveness during the next review cycle.", 5),
            ]
        activity_catalog[practice_code] = activities

practice_rows = []
for objective_code, objective_name, _, _ in objectives:
    for practice_code, practice_name, practice_description, example_metrics in practice_catalog.get(objective_code, []):
        practice_rows.append((objective_code, practice_code, practice_name, practice_description, example_metrics))

activity_rows = []
for _, _, _, _ in objectives:
    pass

for practice_code, activities in activity_catalog.items():
    for activity, level in activities:
        activity_rows.append((practice_code, activity, level))

wb = Workbook()
ws_objectives = wb.active
ws_objectives.title = "objectives"
ws_process = wb.create_sheet("process")
ws_activities = wb.create_sheet("activities")

headers_objectives = ["objective code", "objective name", "objective description", "objective purpose"]
headers_process = ["objective code", "practice code", "practice name", "practice description", "example metrics"]
headers_activities = ["practice code", "activities", "capability level"]

for ws, headers in [(ws_objectives, headers_objectives), (ws_process, headers_process), (ws_activities, headers_activities)]:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

for row in objectives:
    ws_objectives.append(row)

for row in practice_rows:
    ws_process.append(row)

for row in activity_rows:
    ws_activities.append(row)

for ws in [ws_objectives, ws_process, ws_activities]:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

for ws in [ws_objectives, ws_process, ws_activities]:
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)

ws_objectives.freeze_panes = "A2"
ws_process.freeze_panes = "A2"
ws_activities.freeze_panes = "A2"

wb.save(out_path)
print(f"Created {out_path}")
